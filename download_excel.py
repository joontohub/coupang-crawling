import openpyxl


def moveExcel(data,number):

    wb = openpyxl.load_workbook('coupang_exel.xlsx')
    sheet = wb['Sheet1']

    name = data.get('name')
    price = data.get('price') # 이게 실제 판매가격
    ex_price = data.get('ex_price') #이건 할인전 가격
    website = data.get('website')
    print(name, price , ex_price )
    print(website)
    if ex_price != 'None': # 할인전 가격이 있는경우에는 할인전 가격이 정상가격 칸에 들어감
        sheet.cell(row=number+1, column=1).value = name

        ex_price = int(ex_price.replace(',',''))
        sheet.cell(row=number+1, column=2).value = ex_price

        price = int(price.replace(',',''))
        sheet.cell(row=number+1, column=3).value = price

        sheet.cell(row=number+1, column=4).value = website

    else: #할인전가격이 없는 경우에는 실제판매가격이 정상가랑 할인가에 둘 다 들어감
        sheet.cell(row=number+1, column=1).value = name

        price = int(price.replace(',',''))
        sheet.cell(row=number+1, column=2).value = price
        sheet.cell(row=number+1, column=3).value = price

        sheet.cell(row=number+1, column=4).value = website



    wb.save('coupang_exel.xlsx')
